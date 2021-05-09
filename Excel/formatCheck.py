try:
	import openpyxl
	from openpyxl import load_workbook
	from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
	from openpyxl.styles import colors
	from openpyxl.styles import Font, Color
	from openpyxl import Workbook
except:
	print('Import Error')



inputs = Font(color = "0000FF", name = 'Cambria', size = 10)
formulae = Font(color = "000000", name = 'Cambria', size = 10, italic = True)
strings = Font(color = "000000", name = 'Cambria', size = 10)
other = Font(color = "00339966", name = 'Cambria', size = 10)


def formater(x):

	if type(x.value) is str:
	
		if x.value[0:2] == '=\'':
			x.font = other
			return (x.value, 'Other')
			
		elif x.value[0  ] == '=':
			x.font = formulae
			return (x.value, 'Formula')
			
		else:
			x.font = strings
			return (x.value, 'String')
			
	else:
		x.font = inputs
		return (x.value, 'Number')
	

#add check for the titles and underline, make this a copy/paste - run to completion
#of a 2 pager fin statement


dest_filename = './Try.xlsx'
wb = load_workbook(filename = dest_filename)

for sheet in wb.worksheets:
	for row in sheet.iter_rows():
		for cell in row:
			formater(cell)

wb.save(filename = dest_filename)
	
